"""
Impact Analysis Exporter
Generates reports showing how hot list prioritization affects other orders.
"""

import pandas as pd
from datetime import datetime
from typing import List, Dict, Any, Optional
from pathlib import Path


def generate_impact_analysis(
    scheduled_orders: List,
    baseline_orders: List,
    hot_list_entries: List[Dict],
    hot_list_core_shortages: List[Dict],
    output_dir: str
) -> str:
    """
    Generate impact analysis report showing how hot list affects other orders.

    Args:
        scheduled_orders: List of ScheduledOrder objects (with hot list prioritization)
        baseline_orders: List of ScheduledOrder objects (without hot list)
        hot_list_entries: List of hot list entry dictionaries
        hot_list_core_shortages: List of hot list orders that couldn't be scheduled due to missing cores
        output_dir: Output directory path

    Returns:
        Path to the created Excel file
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = output_dir / f"Impact_Analysis_{timestamp}.xlsx"

    # Create lookup for baseline orders by WO#
    baseline_lookup = {o.wo_number: o for o in baseline_orders}

    # Create lookup for hot list WO#s
    hot_list_wo_numbers = {e['wo_number'] for e in hot_list_entries}

    # Analyze delayed orders
    delayed_orders = []
    total_delay_hours = 0
    orders_now_late = 0
    orders_still_on_time = 0

    for order in scheduled_orders:
        if order.wo_number in hot_list_wo_numbers:
            continue  # Skip hot list orders themselves

        baseline = baseline_lookup.get(order.wo_number)
        if not baseline or not baseline.blast_date or not order.blast_date:
            continue

        # Calculate delay in hours
        delay_hours = (order.blast_date - baseline.blast_date).total_seconds() / 3600

        if delay_hours > 0.5:  # Only count delays > 30 minutes
            # Check if on-time status changed
            was_on_time = baseline.on_time
            is_on_time = order.on_time
            status_change = ""
            if was_on_time and not is_on_time:
                status_change = "NOW LATE"
                orders_now_late += 1
            elif is_on_time:
                orders_still_on_time += 1

            delayed_orders.append({
                'WO#': order.wo_number,
                'Serial Number': getattr(order, 'serial_number', None) or '',
                'Part Number': order.part_number,
                'Customer': order.customer,
                'Original BLAST': baseline.blast_date,
                'New BLAST': order.blast_date,
                'Delay (Hours)': round(delay_hours, 1),
                'Delay (Days)': round(delay_hours / 24, 1),
                'Original Completion': baseline.completion_date,
                'New Completion': order.completion_date,
                'Basic Finish Date': order.basic_finish_date,
                'Was On-Time': 'Yes' if was_on_time else 'No',
                'Now On-Time': 'Yes' if is_on_time else 'No',
                'Status Change': status_change
            })
            total_delay_hours += delay_hours

    # Sort by delay (largest first)
    delayed_orders.sort(key=lambda x: -x['Delay (Hours)'])

    # Create summary data
    summary_data = {
        'Metric': [
            'Total Hot List Orders',
            'Hot List ASAP Orders',
            'Hot List Dated Orders',
            'Hot List Core Shortages',
            '',
            'Total Delayed Orders',
            'Average Delay (Hours)',
            'Average Delay (Days)',
            'Total Delay (Hours)',
            '',
            'Orders Now Late (were on-time)',
            'Orders Still On-Time',
        ],
        'Value': [
            len(hot_list_entries),
            sum(1 for e in hot_list_entries if e.get('is_asap')),
            sum(1 for e in hot_list_entries if not e.get('is_asap') and e.get('need_by_date')),
            len(hot_list_core_shortages),
            '',
            len(delayed_orders),
            round(total_delay_hours / len(delayed_orders), 1) if delayed_orders else 0,
            round(total_delay_hours / len(delayed_orders) / 24, 2) if delayed_orders else 0,
            round(total_delay_hours, 1),
            '',
            orders_now_late,
            orders_still_on_time,
        ]
    }

    # Create hot list core shortages data
    shortage_data = []
    for shortage in hot_list_core_shortages:
        entry = shortage.get('hot_list_entry', {})
        shortage_data.append({
            'WO#': shortage.get('wo_number'),
            'Core Needed': shortage.get('core_number_needed'),
            'NEED BY DATE': 'ASAP' if entry.get('is_asap') else entry.get('need_by_date'),
            'DATE REQ MADE': entry.get('date_req_made'),
            'Customer': entry.get('customer'),
            'Description': entry.get('description'),
            'Rubber Override': entry.get('rubber_override'),
        })

    # Write to Excel
    with pd.ExcelWriter(str(output_path), engine='openpyxl') as writer:
        # Summary sheet
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Summary', index=False)

        # Format summary sheet
        ws_summary = writer.sheets['Summary']
        ws_summary.column_dimensions['A'].width = 35
        ws_summary.column_dimensions['B'].width = 15

        # Delayed Orders sheet
        if delayed_orders:
            df_delayed = pd.DataFrame(delayed_orders)
            df_delayed.to_excel(writer, sheet_name='Delayed Orders', index=False)

            # Format delayed orders sheet
            ws_delayed = writer.sheets['Delayed Orders']
            from openpyxl.utils import get_column_letter
            for idx, col in enumerate(df_delayed.columns):
                col_data = df_delayed[col].fillna('').astype(str)
                max_len = max(col_data.str.len().max() if len(col_data) > 0 else 0, len(col)) + 2
                ws_delayed.column_dimensions[get_column_letter(idx + 1)].width = min(max_len, 25)
            ws_delayed.freeze_panes = 'A2'
        else:
            # Create empty sheet with message
            df_empty = pd.DataFrame({'Message': ['No orders were delayed by hot list prioritization']})
            df_empty.to_excel(writer, sheet_name='Delayed Orders', index=False)

        # Hot List Core Shortages sheet
        if shortage_data:
            df_shortages = pd.DataFrame(shortage_data)
            df_shortages.to_excel(writer, sheet_name='Hot List Core Shortages', index=False)

            ws_shortages = writer.sheets['Hot List Core Shortages']
            from openpyxl.utils import get_column_letter
            for idx, col in enumerate(df_shortages.columns):
                col_data = df_shortages[col].fillna('').astype(str)
                max_len = max(col_data.str.len().max() if len(col_data) > 0 else 0, len(col)) + 2
                ws_shortages.column_dimensions[get_column_letter(idx + 1)].width = min(max_len, 25)
            ws_shortages.freeze_panes = 'A2'
        else:
            df_empty = pd.DataFrame({'Message': ['All hot list orders were schedulable']})
            df_empty.to_excel(writer, sheet_name='Hot List Core Shortages', index=False)

    print(f"[OK] Impact analysis exported to: {output_path}")
    return str(output_path)


if __name__ == "__main__":
    print("Impact Analysis Exporter")
    print("This module is designed to be imported and used with scheduler results.")
    print("See excel_exporter.py for usage example.")
