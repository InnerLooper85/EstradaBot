"""
Resource Utilization Exporter
Generates per-station and per-machine utilization metrics from scheduled orders.
"""

import pandas as pd
from collections import defaultdict
from datetime import datetime
from typing import List, Dict, Any


def export_resource_utilization(scheduled_orders: List, output_path: str) -> str:
    """
    Export resource utilization report to Excel.

    Calculates per-station and per-injection-machine utilization from
    the operation history of scheduled orders.

    Args:
        scheduled_orders: List of ScheduledOrder objects with operations
        output_path: Path for output Excel file

    Returns:
        Path to the created file
    """
    if not scheduled_orders:
        # Write an empty report
        pd.DataFrame().to_excel(output_path, index=False)
        return output_path

    # Determine simulation time span from earliest start to latest end
    all_starts = []
    all_ends = []
    for order in scheduled_orders:
        for op in order.operations:
            if op.start_time:
                all_starts.append(op.start_time)
            if op.end_time:
                all_ends.append(op.end_time)

    if not all_starts or not all_ends:
        pd.DataFrame().to_excel(output_path, index=False)
        return output_path

    sim_start = min(all_starts)
    sim_end = max(all_ends)
    total_span_hours = (sim_end - sim_start).total_seconds() / 3600

    if total_span_hours <= 0:
        pd.DataFrame().to_excel(output_path, index=False)
        return output_path

    # --- Station-level utilization ---
    # Group operations by station name
    station_ops = defaultdict(list)
    for order in scheduled_orders:
        for op in order.operations:
            station_ops[op.operation_name].append(op)

    station_rows = []
    for station_name, ops in station_ops.items():
        processing_hours = sum(op.cycle_time for op in ops)
        op_count = len(ops)

        # Sort by start_time to find gaps (idle periods) on this station
        sorted_ops = sorted(ops, key=lambda o: o.start_time)

        # For stations with individual resource_ids (injection machines),
        # we'll handle them separately below — here we aggregate the station total
        station_rows.append({
            'Station': station_name,
            'Resource': '(all)',
            'Orders Processed': op_count,
            'Total Processing Hours': round(processing_hours, 2),
            'Available Hours': round(total_span_hours, 2),
            'Utilization %': round((processing_hours / total_span_hours) * 100, 1),
            'Idle Hours': round(total_span_hours - processing_hours, 2),
        })

    # --- Injection machine-level utilization ---
    machine_ops = defaultdict(list)
    for order in scheduled_orders:
        for op in order.operations:
            if op.resource_id:
                machine_ops[op.resource_id].append(op)

    machine_rows = []
    for machine_id, ops in sorted(machine_ops.items()):
        sorted_ops = sorted(ops, key=lambda o: o.start_time)
        processing_hours = sum(op.cycle_time for op in sorted_ops)
        op_count = len(sorted_ops)

        # Estimate setup/changeover time from gaps between consecutive operations
        # on the same machine (gaps shorter than 2 hours are likely changeover)
        setup_hours = 0.0
        for i in range(1, len(sorted_ops)):
            gap = (sorted_ops[i].start_time - sorted_ops[i - 1].end_time).total_seconds() / 3600
            if 0 < gap <= 2.0:
                setup_hours += gap

        utilized_hours = processing_hours + setup_hours
        idle_hours = total_span_hours - utilized_hours

        machine_rows.append({
            'Station': 'INJECTION',
            'Resource': machine_id,
            'Orders Processed': op_count,
            'Total Processing Hours': round(processing_hours, 2),
            'Setup/Changeover Hours': round(setup_hours, 2),
            'Available Hours': round(total_span_hours, 2),
            'Utilization %': round((utilized_hours / total_span_hours) * 100, 1),
            'Idle Hours': round(max(idle_hours, 0), 2),
        })

    # Build DataFrames
    station_df = pd.DataFrame(station_rows)
    # Sort by utilization descending
    if not station_df.empty:
        station_df = station_df.sort_values('Utilization %', ascending=False)

    machine_df = pd.DataFrame(machine_rows)
    if not machine_df.empty:
        machine_df = machine_df.sort_values('Utilization %', ascending=False)

    # Write to Excel with two sheets
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        station_df.to_excel(writer, sheet_name='Station Utilization', index=False)
        machine_df.to_excel(writer, sheet_name='Machine Utilization', index=False)

        from openpyxl.utils import get_column_letter

        # Format Station Utilization sheet
        ws_station = writer.sheets['Station Utilization']
        for idx, col in enumerate(station_df.columns):
            col_data = station_df[col].fillna('').astype(str)
            max_data_len = col_data.str.len().max() if len(col_data) > 0 else 0
            max_length = max(max_data_len, len(col)) + 2
            ws_station.column_dimensions[get_column_letter(idx + 1)].width = min(max_length, 30)
        ws_station.freeze_panes = 'A2'

        # Format Machine Utilization sheet
        ws_machine = writer.sheets['Machine Utilization']
        for idx, col in enumerate(machine_df.columns):
            col_data = machine_df[col].fillna('').astype(str)
            max_data_len = col_data.str.len().max() if len(col_data) > 0 else 0
            max_length = max(max_data_len, len(col)) + 2
            ws_machine.column_dimensions[get_column_letter(idx + 1)].width = min(max_length, 30)
        ws_machine.freeze_panes = 'A2'

    print(f"[OK] Resource utilization report exported to: {output_path}")
    return output_path
